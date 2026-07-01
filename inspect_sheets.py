import openpyxl
from pathlib import Path

p = Path(r"c:\Users\jhalverstad\Documents\portfolio\html-dataviz-guide\PowerBI_Visual_Catalogus.xlsx")
wb = openpyxl.load_workbook(p)
print("SHEETS:", wb.sheetnames)
for i, name in enumerate(wb.sheetnames, 1):
    ws = wb[name]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    print(f"\n[{i}] {name} | rows={ws.max_row} cols={ws.max_column}")
    print("HEADERS:", headers)
    for r in range(2, min(ws.max_row, 6) + 1):
        vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 8) + 1)]
        print(f"row {r}:", vals)
