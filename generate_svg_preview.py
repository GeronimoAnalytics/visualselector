from pathlib import Path

import openpyxl


WORKBOOK = Path("PowerBI_Visual_Catalogus.xlsx")
OUTPUT_HTML = Path("svg-preview.html")
PRIMARY_SHEET = "PowerBI_Visual_catalogus"
FALLBACK_SHEET = "Visual Catalogus"
SVG_HEADER = "SVG Icoon"


def main() -> None:
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    sheet_name = PRIMARY_SHEET if PRIMARY_SHEET in wb.sheetnames else FALLBACK_SHEET
    ws = wb[sheet_name]

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    if SVG_HEADER not in headers:
        raise ValueError(f"Kolom '{SVG_HEADER}' niet gevonden in sheet '{sheet_name}'.")

    svg_col = headers.index(SVG_HEADER) + 1
    name_col = 1

    cards = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, name_col).value or f"Item {row - 1}"
        svg = ws.cell(row, svg_col).value
        if not svg:
            continue

        cards.append(
            f"""
            <article class=\"card\">
              <div class=\"icon\">{svg}</div>
              <h3>{name}</h3>
            </article>
            """.strip()
        )

    html = f"""<!doctype html>
<html lang=\"nl\">
<head>
  <meta charset=\"utf-8\">
  <meta name=\"viewport\" content=\"width=device-width, initial-scale=1\">
  <title>SVG Icoon Preview</title>
  <style>
    :root {{
      --bg: #eef6fb;
      --panel: #ffffff;
      --text: #17374a;
      --muted: #5c7a8b;
      --line: #c3dbe8;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: "Segoe UI", Arial, sans-serif;
      color: var(--text);
      background: radial-gradient(circle at top right, #d8eef9, var(--bg));
    }}
    header {{
      padding: 20px 22px 12px;
    }}
    h1 {{
      margin: 0;
      font-size: 24px;
    }}
    p {{
      margin: 6px 0 0;
      color: var(--muted);
    }}
    main {{
      padding: 10px 20px 24px;
      display: grid;
      gap: 12px;
      grid-template-columns: repeat(auto-fill, minmax(210px, 1fr));
    }}
    .card {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 14px;
      padding: 12px;
      box-shadow: 0 4px 14px rgba(16, 74, 110, 0.08);
      display: grid;
      gap: 10px;
    }}
    .icon {{
      border: 1px solid var(--line);
      border-radius: 10px;
      background: #f6fbff;
      padding: 8px;
      min-height: 88px;
      display: grid;
      place-items: center;
    }}
    .icon svg {{
      width: 100%;
      height: auto;
      max-height: 74px;
    }}
    h3 {{
      margin: 0;
      font-size: 14px;
      line-height: 1.3;
      font-weight: 600;
    }}
  </style>
</head>
<body>
  <header>
    <h1>SVG Icoon Preview</h1>
    <p>{len(cards)} iconen uit '{sheet_name}' in '{WORKBOOK.name}'.</p>
  </header>
  <main>
    {''.join(cards)}
  </main>
</body>
</html>
"""

    OUTPUT_HTML.write_text(html, encoding="utf-8")
    print(f"Preview gegenereerd: {OUTPUT_HTML.resolve()}")


if __name__ == "__main__":
    main()