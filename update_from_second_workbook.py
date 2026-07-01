import openpyxl
from pathlib import Path

base = Path(r"c:\Users\jhalverstad\Documents\portfolio\html-dataviz-guide")
target_path = base / "PowerBI_Visual_Catalogus.xlsx"
source_path = base / "PowerBI_Visual_Catalogus2.xlsx"

wb_t = openpyxl.load_workbook(target_path)
wb_s = openpyxl.load_workbook(source_path, data_only=True)

ws_t = wb_t["Visual Catalogus"]
ws_s = wb_s["Visual Catalogus"]

# Target headers on row 1
t_headers = [ws_t.cell(1, c).value for c in range(1, ws_t.max_column + 1)]
t_idx = {h: i + 1 for i, h in enumerate(t_headers) if h}

# Source header row is not row 1; detect by the row containing "Naam"
s_header_row = None
for r in range(1, 15):
    row_vals = [ws_s.cell(r, c).value for c in range(1, ws_s.max_column + 1)]
    if "Naam" in row_vals:
        s_header_row = r
        break

if not s_header_row:
    raise ValueError("Kon header rij in bron niet vinden.")

s_headers = [ws_s.cell(s_header_row, c).value for c in range(1, ws_s.max_column + 1)]
s_idx = {h: i + 1 for i, h in enumerate(s_headers) if h}

if "Naam" not in s_idx:
    raise ValueError("Bron mist kolom Naam.")

source_rows_by_name = {}
for r in range(s_header_row + 1, ws_s.max_row + 1):
    name = ws_s.cell(r, s_idx["Naam"]).value
    if name:
        source_rows_by_name[str(name).strip()] = r

copied = 0
missing = []

# Copy shared columns except SVG Icoon (stays in target)
shared = [h for h in t_idx if h in s_idx and h != "SVG Icoon"]
for r in range(2, ws_t.max_row + 1):
    name = ws_t.cell(r, t_idx["Naam"]).value
    if not name:
        continue
    key = str(name).strip()
    sr = source_rows_by_name.get(key)
    if not sr:
        missing.append(key)
        continue

    for h in shared:
        ws_t.cell(r, t_idx[h]).value = ws_s.cell(sr, s_idx[h]).value
    copied += 1

# Remove "Aanbevolen Voor" column from target
if "Aanbevolen Voor" in t_idx:
    ws_t.delete_cols(t_idx["Aanbevolen Voor"], 1)

wb_t.save(target_path)

print(f"Updated rows from source: {copied}")
print(f"Missing names not found in source: {len(missing)}")
if missing:
    print("First 10 missing:", missing[:10])
print("Done. Column 'Aanbevolen Voor' removed if present.")
