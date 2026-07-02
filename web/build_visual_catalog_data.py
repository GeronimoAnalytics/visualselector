import json
from pathlib import Path

import openpyxl

WORKBOOK = Path(__file__).resolve().parents[1] / "PowerBI_Visual_Catalogus.xlsx"
OUTPUT = Path(__file__).resolve().parent / "visuals.json"
OUTPUT_JS = Path(__file__).resolve().parent / "visuals-data.js"

SHEET_CANDIDATES = ["PowerBI_Visual_catalogus", "Visual Catalogus"]

GOAL_MAP = {
    "vergelijking": "comparison",
    "trend": "time-series",
    "samenstelling": "part-to-whole",
    "relatie": "relationship",
    "geografisch": "map",
    "proces": "flow",
    "detail": "detail",
    "kpi": "kpi",
    "analyse": "analysis",
    "combinatie": "mixed",
}

ORDER_HINTS = {
    "comparison": "unordered",
    "time-series": "ordered",
    "part-to-whole": "unordered",
    "relationship": "unordered",
    "map": "unordered",
    "flow": "ordered",
    "detail": "both",
    "kpi": "ordered",
    "analysis": "ordered",
    "mixed": "both",
}

DATA_TYPE_HINTS = {
    "comparison": "both",
    "time-series": "numeric",
    "part-to-whole": "both",
    "relationship": "numeric",
    "map": "both",
    "flow": "both",
    "detail": "both",
    "kpi": "numeric",
    "analysis": "numeric",
    "mixed": "both",
}

ICON_PATH_HEADERS = [
    "Icoon Pad",
    "Icon Pad",
    "Icon Path",
    "SVG Pad",
    "SVG Path",
    "SVG Icoon Pad",
]


def normalize_slug(value: str) -> str:
    return "".join(ch.lower() if ch.isalnum() else "-" for ch in value).strip("-")


def first_existing_sheet(workbook):
    for name in SHEET_CANDIDATES:
        if name in workbook.sheetnames:
            return workbook[name]
    raise ValueError(f"Geen van deze sheets gevonden: {SHEET_CANDIDATES}")


def pick_first_column(index: dict[str, int], candidates: list[str]) -> int | None:
    for candidate in candidates:
        if candidate in index:
            return index[candidate]
    return None


def main() -> None:
    workbook = openpyxl.load_workbook(WORKBOOK, data_only=True)
    sheet = first_existing_sheet(workbook)

    headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    index = {header: idx + 1 for idx, header in enumerate(headers) if header}

    required = [
        "Naam",
        "Soort",
        "Categorie",
        "Complexiteit",
        "Wat is het?",
        "Wanneer gebruiken?",
        "Wanneer niet gebruiken?",
        "Afbeelding Pad",
    ]
    missing = [key for key in required if key not in index]
    if missing:
        raise ValueError(f"Vereiste kolommen missen: {missing}")

    svg_icon_col = index.get("SVG Icoon")
    icon_path_col = pick_first_column(index, ICON_PATH_HEADERS)
    icons_dir = WORKBOOK.parent / "icons"

    visuals = []
    for row in range(2, sheet.max_row + 1):
        name = sheet.cell(row, index["Naam"]).value
        if not name:
            continue

        category = (sheet.cell(row, index["Categorie"]).value or "").strip()
        category_key = category.lower()
        goal = GOAL_MAP.get(category_key, "mixed")

        visual_id = normalize_slug(str(name))
        icon_path = ""

        if icon_path_col:
            icon_path = sheet.cell(row, icon_path_col).value or ""
            icon_path = str(icon_path).strip()

        if not icon_path:
            candidate = icons_dir / f"{visual_id}.svg"
            if candidate.exists():
                icon_path = f"icons/{candidate.name}"

        visuals.append(
            {
                "id": visual_id,
                "name": str(name),
                "kind": sheet.cell(row, index["Soort"]).value or "Visual",
                "category": category,
                "goal": goal,
                "complexity": sheet.cell(row, index["Complexiteit"]).value or "Onbekend",
                "whatIsIt": sheet.cell(row, index["Wat is het?"]).value or "",
                "whenToUse": sheet.cell(row, index["Wanneer gebruiken?"]).value or "",
                "whenNotToUse": sheet.cell(row, index["Wanneer niet gebruiken?"]).value or "",
                "imagePath": sheet.cell(row, index["Afbeelding Pad"]).value or "",
                "iconPath": icon_path,
                "svgIcon": sheet.cell(row, svg_icon_col).value or "" if svg_icon_col else "",
                "orderType": ORDER_HINTS.get(goal, "both"),
                "dataType": DATA_TYPE_HINTS.get(goal, "both"),
            }
        )

    payload = {
        "meta": {
            "title": "Power BI Visual Catalogus",
            "source": WORKBOOK.name,
            "sheet": sheet.title,
            "count": len(visuals),
        },
        "tooltips": {
            "ordered": {
                "title": "Ordered data",
                "description": "Data met natuurlijke volgorde, zoals tijd of schaal.",
                "examples": ["Januari t/m december", "Leeftijd 18, 19, 20", "Fase 1 -> 2 -> 3"],
            },
            "unordered": {
                "title": "Unordered data",
                "description": "Categorieen zonder vaste volgorde.",
                "examples": ["Product A, B, C", "Regio Noord/Zuid", "Teams Sales/Finance/IT"],
            },
            "numeric": {
                "title": "Numeric",
                "description": "Waarden die je kunt optellen, middelen of vergelijken.",
                "examples": ["Omzet", "Marge %", "Aantal klanten"],
            },
            "categoric": {
                "title": "Categoric",
                "description": "Labels of groepen die data indelen.",
                "examples": ["Land", "Segment", "Klanttype"],
            },
        },
        "visuals": visuals,
    }

    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    OUTPUT_JS.write_text(
        "window.CATALOG_DATA = " + json.dumps(payload, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )
    print(f"Geschreven: {OUTPUT} ({len(visuals)} visuals)")
    print(f"Geschreven: {OUTPUT_JS}")


if __name__ == "__main__":
    main()
