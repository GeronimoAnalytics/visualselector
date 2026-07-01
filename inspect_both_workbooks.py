import openpyxl
from pathlib import Path

for fn in ["PowerBI_Visual_Catalogus.xlsx", "PowerBI_Visual_Catalogus2.xlsx"]:
    p = Path(r"c:\Users\jhalverstad\Documents\portfolio\html-dataviz-guide") / fn
    wb = openpyxl.load_workbook(p)
    print("\nFILE:", fn)
    print("SHEETS:", wb.sheetnames)
    ws = wb[wb.sheetnames[0]]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    print("HEADERS:", headers)
    for r in range(2, min(ws.max_row, 8) + 1):
        vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 9) + 1)]
        print(f"row {r}:", vals)
