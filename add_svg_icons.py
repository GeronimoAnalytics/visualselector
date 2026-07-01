import openpyxl
from pathlib import Path

workbook_path = Path(r"c:\Users\jhalverstad\Documents\portfolio\html-dataviz-guide\PowerBI_Visual_Catalogus.xlsx")
wb = openpyxl.load_workbook(workbook_path)

sheet_name = "PowerBI_Visual_catalogus" if "PowerBI_Visual_catalogus" in wb.sheetnames else "Visual Catalogus"
ws = wb[sheet_name]

HEADER = "SVG Icoon"
if ws.cell(1, ws.max_column).value != HEADER:
    target_col = ws.max_column + 1
    ws.cell(1, target_col).value = HEADER
else:
    target_col = ws.max_column

# Palette inspired by the provided catalog image.
C1 = "#0B4F8A"
C2 = "#1492C5"
C3 = "#F2C24D"
BG1 = "#EAF6FF"
BG2 = "#CFE8F7"
GRID = "#9BC2DA"


def wrap(inner: str) -> str:
    return (
        "<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 96 64'>"
        "<defs><linearGradient id='bg' x1='0' y1='0' x2='0' y2='1'>"
        f"<stop offset='0%' stop-color='{BG1}'/><stop offset='100%' stop-color='{BG2}'/>"
        "</linearGradient></defs>"
        "<rect x='1' y='1' width='94' height='62' rx='10' fill='url(#bg)' stroke='#7EA9C3'/>"
        f"{inner}</svg>"
    )


def icon_bar() -> str:
    return wrap(
        f"<rect x='12' y='14' width='30' height='8' fill='{C1}'/>"
        f"<rect x='12' y='27' width='46' height='8' fill='{C2}'/>"
        f"<rect x='12' y='40' width='62' height='8' fill='{C3}'/>"
    )


def icon_column() -> str:
    return wrap(
        f"<rect x='18' y='30' width='12' height='20' fill='{C1}'/>"
        f"<rect x='36' y='20' width='12' height='30' fill='{C2}'/>"
        f"<rect x='54' y='12' width='12' height='38' fill='{C3}'/>"
        "<line x1='12' y1='50' x2='82' y2='50' stroke='#4F7A94'/>"
    )


def icon_line() -> str:
    return wrap(
        "<polyline points='14,44 30,34 44,39 58,24 76,16' fill='none' stroke='#0B4F8A' stroke-width='3'/>"
        "<circle cx='30' cy='34' r='2.5' fill='#1492C5'/><circle cx='58' cy='24' r='2.5' fill='#F2C24D'/>"
    )


def icon_area(stacked: bool = False) -> str:
    if stacked:
        return wrap(
            f"<path d='M12 50 L12 40 L28 33 L46 38 L64 28 L82 34 L82 50 Z' fill='{C1}' opacity='0.75'/>"
            f"<path d='M12 40 L28 30 L46 33 L64 22 L82 26 L82 34 L64 28 L46 38 L28 33 L12 40 Z' fill='{C2}' opacity='0.75'/>"
            f"<path d='M12 30 L28 20 L46 24 L64 14 L82 17 L82 26 L64 22 L46 33 L28 30 L12 40 Z' fill='{C3}' opacity='0.75'/>"
        )
    return wrap(
        f"<path d='M12 50 L12 38 L30 28 L46 31 L64 19 L82 16 L82 50 Z' fill='{C2}' opacity='0.9'/>"
        "<polyline points='12,38 30,28 46,31 64,19 82,16' fill='none' stroke='#0B4F8A' stroke-width='2'/>"
    )


def icon_ribbon() -> str:
    return wrap(
        f"<path d='M12 20 C28 10, 40 32, 54 22 C66 14, 76 34, 84 24 L84 34 C74 44, 66 24, 54 32 C40 42, 26 18, 12 30 Z' fill='{C2}'/>"
        f"<path d='M12 30 C26 18, 40 42, 54 32 C66 24, 74 44, 84 34 L84 42 C74 52, 66 32, 54 40 C40 50, 28 28, 12 38 Z' fill='{C3}' opacity='0.9'/>"
    )


def icon_waterfall() -> str:
    return wrap(
        "<line x1='12' y1='50' x2='82' y2='50' stroke='#4F7A94'/>"
        f"<rect x='16' y='34' width='10' height='16' fill='{C1}'/>"
        f"<rect x='32' y='24' width='10' height='10' fill='{C2}'/>"
        f"<rect x='48' y='24' width='10' height='18' fill='{C3}'/>"
        f"<rect x='64' y='42' width='10' height='8' fill='{C1}'/>"
        "<line x1='26' y1='34' x2='32' y2='34' stroke='#4F7A94'/><line x1='42' y1='24' x2='48' y2='24' stroke='#4F7A94'/>"
    )


def icon_pie(donut: bool = False) -> str:
    hole = "<circle cx='48' cy='32' r='10' fill='url(#bg)'/>" if donut else ""
    return wrap(
        "<circle cx='48' cy='32' r='20' fill='#0B4F8A'/>"
        "<path d='M48 32 L48 12 A20 20 0 0 1 66 44 Z' fill='#1492C5'/>"
        "<path d='M48 32 L66 44 A20 20 0 0 1 32 46 Z' fill='#F2C24D'/>"
        f"{hole}"
    )


def icon_treemap() -> str:
    return wrap(
        f"<rect x='14' y='14' width='30' height='22' fill='{C1}'/>"
        f"<rect x='46' y='14' width='36' height='14' fill='{C2}'/>"
        f"<rect x='46' y='30' width='16' height='22' fill='{C3}'/>"
        f"<rect x='64' y='30' width='18' height='10' fill='{C1}' opacity='0.8'/><rect x='64' y='42' width='18' height='10' fill='{C2}' opacity='0.8'/>"
    )


def icon_funnel() -> str:
    return wrap(
        f"<path d='M16 14 H80 L66 28 H30 Z' fill='{C1}'/>"
        f"<path d='M30 28 H66 L58 40 H38 Z' fill='{C2}'/>"
        f"<path d='M38 40 H58 L52 50 H44 Z' fill='{C3}'/>"
    )


def icon_scatter(bubble: bool = False) -> str:
    if bubble:
        return wrap(
            "<line x1='14' y1='50' x2='82' y2='50' stroke='#4F7A94'/><line x1='14' y1='50' x2='14' y2='14' stroke='#4F7A94'/>"
            f"<circle cx='28' cy='40' r='4' fill='{C1}'/><circle cx='44' cy='32' r='6' fill='{C2}'/><circle cx='62' cy='24' r='8' fill='{C3}'/>"
        )
    return wrap(
        "<line x1='14' y1='50' x2='82' y2='50' stroke='#4F7A94'/><line x1='14' y1='50' x2='14' y2='14' stroke='#4F7A94'/>"
        f"<circle cx='26' cy='42' r='2.5' fill='{C1}'/><circle cx='35' cy='35' r='2.5' fill='{C2}'/><circle cx='46' cy='30' r='2.5' fill='{C3}'/><circle cx='60' cy='24' r='2.5' fill='{C1}'/><circle cx='70' cy='20' r='2.5' fill='{C2}'/>"
    )


def icon_table(matrix: bool = False) -> str:
    grid = ""
    for x in (14, 34, 54, 74):
        grid += f"<line x1='{x}' y1='16' x2='{x}' y2='50' stroke='{GRID}'/>"
    for y in (16, 27, 38, 50):
        grid += f"<line x1='14' y1='{y}' x2='74' y2='{y}' stroke='{GRID}'/>"
    accent = f"<rect x='14' y='16' width='60' height='11' fill='{C1}' opacity='0.25'/>" if matrix else f"<rect x='14' y='16' width='20' height='34' fill='{C2}' opacity='0.15'/>"
    return wrap(accent + grid)


def icon_card(multi: bool = False, kpi: bool = False, gauge: bool = False) -> str:
    if gauge:
        return wrap(
            "<path d='M24 42 A24 24 0 0 1 72 42' fill='none' stroke='#0B4F8A' stroke-width='6'/>"
            "<path d='M24 42 A24 24 0 0 1 55 20' fill='none' stroke='#1492C5' stroke-width='6'/>"
            "<line x1='48' y1='42' x2='64' y2='28' stroke='#F2C24D' stroke-width='3'/>"
        )
    if kpi:
        return wrap(
            f"<rect x='14' y='14' width='42' height='36' rx='6' fill='white' stroke='{GRID}'/>"
            f"<polyline points='20,40 28,34 36,36 46,24' fill='none' stroke='{C2}' stroke-width='2'/>"
            f"<circle cx='68' cy='32' r='12' fill='{C3}'/><path d='M62 33 L67 38 L75 27' fill='none' stroke='#1C5E2A' stroke-width='3'/>"
        )
    if multi:
        return wrap(
            f"<rect x='14' y='16' width='68' height='10' rx='3' fill='{C1}' opacity='0.25'/>"
            f"<rect x='14' y='30' width='54' height='8' rx='3' fill='{C2}' opacity='0.35'/>"
            f"<rect x='14' y='42' width='40' height='8' rx='3' fill='{C3}' opacity='0.5'/>"
        )
    return wrap(
        "<rect x='14' y='14' width='68' height='36' rx='6' fill='white' stroke='#9BC2DA'/>"
        "<text x='48' y='37' text-anchor='middle' font-size='14' font-family='Segoe UI, Arial' fill='#0B4F8A'>1.2M</text>"
    )


def icon_map(filled: bool = False) -> str:
    if filled:
        return wrap(
            f"<path d='M24 20 L40 14 L56 20 L70 16 L72 44 L54 50 L36 46 L22 50 Z' fill='{C2}' opacity='0.8' stroke='{C1}'/>"
            f"<path d='M40 14 L40 46 M56 20 L54 50 M24 20 L36 46 M70 16 L72 44' stroke='white' opacity='0.5'/>"
        )
    return wrap(
        f"<circle cx='48' cy='32' r='20' fill='{C2}' opacity='0.3' stroke='{C1}'/>"
        f"<path d='M38 22 L46 18 L56 22 L58 30 L52 38 L44 36 L38 28 Z' fill='{C1}' opacity='0.8'/>"
        f"<circle cx='66' cy='24' r='3' fill='{C3}'/><circle cx='30' cy='40' r='3' fill='{C3}'/>"
    )


def icon_tree() -> str:
    return wrap(
        f"<rect x='14' y='18' width='16' height='8' fill='{C1}'/><rect x='40' y='12' width='16' height='8' fill='{C2}'/><rect x='40' y='26' width='16' height='8' fill='{C2}'/><rect x='66' y='8' width='16' height='8' fill='{C3}'/><rect x='66' y='20' width='16' height='8' fill='{C3}'/><rect x='66' y='32' width='16' height='8' fill='{C3}'/>"
        "<path d='M30 22 H40 M56 16 H66 M56 30 H66 M56 22 H66' stroke='#4F7A94'/>"
    )


def icon_text() -> str:
    return wrap(
        "<rect x='14' y='14' width='68' height='36' rx='5' fill='white' stroke='#9BC2DA'/>"
        "<line x1='20' y1='24' x2='74' y2='24' stroke='#0B4F8A'/><line x1='20' y1='31' x2='70' y2='31' stroke='#1492C5'/><line x1='20' y1='38' x2='62' y2='38' stroke='#1492C5'/>"
    )


def icon_button() -> str:
    return wrap(
        f"<rect x='24' y='20' width='48' height='24' rx='12' fill='{C1}'/>"
        "<text x='48' y='35' text-anchor='middle' font-size='11' font-family='Segoe UI, Arial' fill='white'>GO</text>"
    )


def icon_slicer() -> str:
    return wrap(
        "<rect x='16' y='14' width='64' height='36' rx='6' fill='white' stroke='#9BC2DA'/>"
        f"<rect x='22' y='20' width='52' height='8' rx='3' fill='{C2}' opacity='0.35'/>"
        f"<rect x='22' y='32' width='38' height='8' rx='3' fill='{C1}' opacity='0.35'/>"
        f"<circle cx='64' cy='36' r='4' fill='{C3}'/>"
    )


def icon_generic() -> str:
    return wrap(
        f"<rect x='18' y='18' width='16' height='28' fill='{C1}'/><rect x='40' y='24' width='16' height='22' fill='{C2}'/><rect x='62' y='14' width='16' height='32' fill='{C3}'/>"
    )


def icon_for(name: str, category: str) -> str:
    n = (name or "").lower()
    c = (category or "").lower()

    if "bar" in n:
        return icon_bar()
    if "column" in n:
        return icon_column()
    if "line and" in n:
        return wrap(
            f"<rect x='16' y='28' width='10' height='20' fill='{C1}'/><rect x='30' y='22' width='10' height='26' fill='{C2}'/><rect x='44' y='16' width='10' height='32' fill='{C3}'/>"
            "<polyline points='16,18 34,24 48,20 64,14 78,16' fill='none' stroke='#0B4F8A' stroke-width='2.5'/>"
        )
    if "line chart" in n:
        return icon_line()
    if "stacked area" in n:
        return icon_area(stacked=True)
    if "area" in n:
        return icon_area()
    if "ribbon" in n:
        return icon_ribbon()
    if "waterfall" in n:
        return icon_waterfall()
    if "pie" in n:
        return icon_pie()
    if "donut" in n:
        return icon_pie(donut=True)
    if "treemap" in n:
        return icon_treemap()
    if "funnel" in n:
        return icon_funnel()
    if "bubble" in n:
        return icon_scatter(bubble=True)
    if "scatter" in n or "dot plot" in n:
        return icon_scatter()
    if "matrix" in n:
        return icon_table(matrix=True)
    if n == "table":
        return icon_table()
    if "kpi" in n or "goals" in n or "metrics" in n:
        return icon_card(kpi=True)
    if "multi-row" in n:
        return icon_card(multi=True)
    if "gauge" in n:
        return icon_card(gauge=True)
    if "card" in n:
        return icon_card()
    if "map" in n or "arcgis" in n or "azure" in n:
        return icon_map(filled=("filled" in n or "shape" in n))
    if "decomposition" in n or "influencers" in n:
        return icon_tree()
    if "narrative" in n or "q&a" in n or "text" in n:
        return icon_text()
    if "image" in n:
        return wrap(
            f"<rect x='16' y='14' width='64' height='36' rx='6' fill='white' stroke='{GRID}'/>"
            f"<path d='M24 42 L38 28 L48 38 L58 26 L72 42 Z' fill='{C2}' opacity='0.6'/><circle cx='32' cy='24' r='4' fill='{C3}'/>"
        )
    if "shape" in n:
        return wrap(
            f"<circle cx='30' cy='32' r='10' fill='{C2}'/><rect x='44' y='22' width='18' height='18' fill='{C1}'/><polygon points='74,42 64,22 84,22' fill='{C3}'/>"
        )
    if "button" in n or "navigator" in n:
        return icon_button()
    if "power apps" in n or "power automate" in n or "paginated" in n:
        return wrap(
            f"<rect x='16' y='14' width='22' height='36' rx='5' fill='{C1}'/><rect x='42' y='14' width='16' height='36' rx='5' fill='{C2}'/><rect x='62' y='14' width='18' height='36' rx='5' fill='{C3}'/>"
        )
    if "python" in n or "r visual" in n:
        return wrap(
            "<text x='31' y='38' text-anchor='middle' font-size='22' font-family='Segoe UI, Arial' fill='#0B4F8A'>Py</text>"
            "<text x='64' y='38' text-anchor='middle' font-size='22' font-family='Segoe UI, Arial' fill='#1492C5'>R</text>"
        )
    if "slicer" in n:
        return icon_slicer()
    if any(k in n for k in ["forecast", "anomaly", "error", "average", "constant", "median", "min", "max", "percentile", "analytics"]):
        return icon_line()
    if any(k in n for k in ["tooltip", "drillthrough", "cross", "bookmarks", "selection pane", "sync", "personalize", "visual calculations", "quick measures", "field parameters", "calculation groups"]):
        return icon_tree()
    if "small multiples" in n:
        return wrap(
            f"<rect x='14' y='16' width='28' height='16' fill='{C1}' opacity='0.65'/><rect x='46' y='16' width='28' height='16' fill='{C2}' opacity='0.65'/><rect x='14' y='34' width='28' height='16' fill='{C2}' opacity='0.65'/><rect x='46' y='34' width='28' height='16' fill='{C3}' opacity='0.65'/>"
        )
    if any(k in n for k in ["hierarchical slicer", "relative date", "relative time", "dropdown slicer", "list slicer", "between slicer"]):
        return icon_slicer()

    if c in {"vergelijking", "trend", "samenstelling", "kpi", "geografisch", "relatie"}:
        return {
            "vergelijking": icon_column,
            "trend": icon_line,
            "samenstelling": icon_treemap,
            "kpi": lambda: icon_card(kpi=True),
            "geografisch": icon_map,
            "relatie": icon_scatter,
        }[c]()

    return icon_generic()


for row in range(2, ws.max_row + 1):
    name = ws.cell(row=row, column=1).value
    category = ws.cell(row=row, column=3).value
    ws.cell(row=row, column=target_col).value = icon_for(name, category)

ws.column_dimensions[openpyxl.utils.get_column_letter(target_col)].width = 80
wb.save(workbook_path)
print(f"Updated sheet '{sheet_name}' with SVG icons in column {target_col} ({HEADER}).")
